import openpyxl, re, sys, json

def clean(s):
    return re.sub(r'\s+', ' ', re.sub(r'[^a-z0-9 ]', ' ', s.lower().strip())).strip()

our_services = [
    'Echocardiography','ECG','Delivery','Anti Natal','Abdomino Pelvic Scan','Pelvic Scan',
    'Obstetric Scan','Wound Dressing','Susp. Priton','Syrup M. Vite','Syrup PCM',
    'Syrup Emzolyn Child IBT','Syrup Emzolyn Adult IBT','Syrup Vit C','Susp. Fluconazole',
    'Susp. Metro','Susp Ibuprofen','Susp Lactulose','Susp. Amoxiclav','Susp. Albendazole',
    'Susp. Erythromycin','Susp. Loratadine','Susp. Artemether','Susp. Amoxil 125mg','Susp Asad',
    'Acceclofenac 100mg','Diclofenac 100mg','Tab Erythromycin','Tab Levofloxacin','Tab Lexotan X10',
    'Tab Misoprostol','Tab Metro 200mg X1','Tab Vaginal Insert X1','Tab PCM 500mg X18',
    'Tab Acycorplus','Tab Omeprazole 20mg X14','Tab Ibuprofen 400mg','Tab Cefurozine 500mg X10',
    'Tab Arthemeter 80/480mg X6','Tab Ciprofloxacine 500mg X10','Tab Aldomet X10',
    'Tab Vasoprin 75mg X10','Tab Daonil 5mg X10','Tab Hyperex 20mg SR X10','Tab Metformin 500mg X10',
    'Tab Atenolol 100mg X14','Tab Lisinooril 10mg X14','Tab Amlodipine 10mg X14','Tab Propranol X40mg',
    'Tab Moderatic X10','Tab Laxis 40mg X10','Tab Losartan Potassium X14','Tab Ibuprofen 400mg X10',
    'Tab Buscopam 10mg X10','Tab Dexamethasone X10','Tab Tinidazole 500mg X10','Tab Ampiclox 500mg X10',
    'Tab Bendro Fluazide 5mg','Tab Prednisalone','Tab ACT 3','Tab ACT 2','Tab Ibasunate',
    'Tab Pregabalin 75mg','Tab Arthrotec','Zadol Forte','Tab Augmentin','Tab Clarithromycin',
    'Tab Azithromycin 500mg','Tab Ketoconazole','Tab Bromazepan 3mg','Tab Cotrimozole 1 X1',
    'Tab Clopidogrel 75mg','Tab Cocodamol','Tab Fansidar X3','Tab Fasolate','Tab Folic Acid',
    'Tab Daonil/Gilbenclamide','Tab Vit B. Complex','Tab Tenofovir Combination','Tab No-ACH',
    'Tab Sirdalude 4mg','Tab Salbutamol 4mg','Tab Ofloxacin','Tab Amitripphyline','Tab Zinc Sulphate',
    'Tab Vit C X1','Tab Loperamide','Tab Cemetadine 400mg','Tab Dana (Gelosyl)','Tab Loratidine',
    'Tab Nifedifine 20mg','Tab Vit E','Tab Vit A','Tab Livolin Forte Card','Blood Tonic',
    'Cap Fluconazole 200mg','Cap Amoxil Adult','Amoxil Child','Mirapain Cream','Clotrimazole Cream',
    'Clotrimazole Insert','Neurogestic Cream','Ketoconazole Cream','Neoskin Cream','Diclofenac Cream',
    'Scale Vein','Given Set','Blood Giving Set','IV Canula','Catheter','Urine Bag','Surgical Glove',
    'Gauze','Bandage','Syringe 2mls','Syringes 5mls','Syringes 10mls','ORS X1','N. G Tube',
    'Inj. Penta','Inj. Hyosin','Inj. PCM 300mg','Inj. Diazepan','Inj. Diclo 75mg','Inj. Hydrocort',
    'Inj. Omeprazole 70mg','IV Metro','IV Cipro','IV Omeprazole','Inj. Insulin','Inj. Genta',
    'Inj. Exacef','Inj. Dexamethasone','Inj. Plasil','Inj. Prometh','Inj. Frusamide','Inj. Quinine',
    'Inj. Ranitidine','IV Artesunate 120mg','Inj. Artemether 80mg X6','Inj. Imal Non Branded',
    'Inj. Imal Branded','Inj. Aminophyllin','Inj. Tandak','IV Ceftriaxone LG',
    'IV Artesunate 60mg/Rekmal','IV Rekmal 30mg','Inj T.T','Inj. B. Complex',
    'Inj. Amoxiclav 1.2mg','Inj. Amoxiclav 0.6','IVF D/S','IVF N/S','IVF D/W','IVF 4.3',
    'IVF Ringers Lactate','IVF Manitol','IV FSD','Piroxicam','Doxycycline','Fluconazole',
    'Calamine Lotion','Neurovite Forte','Antacid Susp','Artemethine Lumephantone 80/486',
    'MP (Malaria Parasite)','Widal','PCV','Urinalysis','PT (Pregnancy Test)','HCV','HBV',
    'ARC/RVS','FBC (Full Blood Count)','ESR','Blood Grouping','RBS (Random Blood Sugar)',
    'FBS (Fasting Blood Sugar)','2HPP','Urine M/C/S','Stool M/C/S','Stool Microscopy',
    'Occult Blood','HVS MCS','HbA1C','AFP ()','Lipid Profile','Genotype','PSA Qualitative',
    'PSA Quantitative','Rheumatoid Factor','H. Pylori','Urea','Creatinine','Cholesterol',
    'Uric Acid','Calcium','ASL','ALT','LFT (Liver Function Test)','E/U/CR','Electrolyte',
    'Albumin','HBV Quantification','HBV (DNA)','HCV RNA','BCV Genotype',
    'Viral Screening (Donor Screening)','T3 T4 and TSA','HB Serology/Profile','Total Bilirubin',
    'Plural Fluid AFB','RFT/E/U/Cr','VDRL','ARC Viral Load (RNA)','Asetic Fluid Analysis (AFAB)',
    'AFAC','AFA Haem','A Feto Protein Quantitative','Magnesium','Organic Phosphate','Viral Maker',
    'Rheumatoid Factor +VE','Blood Culture','CD4 Count',
]
our_clean = {clean(s): s for s in our_services}

wb = openpyxl.load_workbook('UPDATED AVON -- TARIFF.xlsx')
ws = wb['ALL'] if 'ALL' in wb.sheetnames else wb.active

matched = []
unmatched = []
seen = set()

for row in ws.iter_rows(values_only=True):
    if not row[1] or not row[3] or not isinstance(row[3], (int, float)):
        continue
    svc_name = str(row[1]).strip()
    price = float(row[3])
    level = str(row[2]).strip() if row[2] else ''
    tc = clean(svc_name)
    if svc_name in seen:
        continue
    seen.add(svc_name)

    # Category from sheet context (col 2 = LEVEL not sheet)
    cat = 'Pharmacy'

    found = None
    if tc in our_clean:
        found = our_clean[tc]
    else:
        for oc, oname in our_clean.items():
            oc_w = set(oc.split())
            tc_w = set(tc.split())
            if len(oc_w) >= 2 and oc_w.issubset(tc_w):
                found = oname
                break
            elif len(oc_w) == 1 and list(oc_w)[0] in tc_w and len(tc_w) <= 3:
                found = oname
                break

    if found:
        matched.append((found, price))
    else:
        unmatched.append((svc_name, cat, price))

print(f'Matched: {len(matched)}  Unmatched: {len(unmatched)}')
with open('avoh_matched.json', 'w', encoding='utf-8') as f:
    json.dump(matched, f, ensure_ascii=False)
with open('avoh_unmatched.json', 'w', encoding='utf-8') as f:
    json.dump(unmatched, f, ensure_ascii=False)
print('Done. Check avoh_matched.json and avoh_unmatched.json')
